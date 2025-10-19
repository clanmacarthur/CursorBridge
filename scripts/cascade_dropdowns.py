import argparse
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook  # type: ignore
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore


def sanitize_key(text: str) -> str:
    key = re.sub(r"[^A-Za-z0-9]", "_", text.strip())
    if not key or not key[0].isalpha():
        key = f"N_{key}" if key else "N_"
    return re.sub(r"_+", "_", key)


def headers_of(ws) -> List[str]:
    return [str(c.value).strip() if c.value is not None else "" for c in ws[1]]


def find_col(headers: List[str], candidates: List[str]) -> Optional[int]:
    lookup = {(h or "").strip().lower(): i for i, h in enumerate(headers, start=1)}
    for cand in candidates:
        idx = lookup.get(cand.strip().lower())
        if idx is not None:
            return idx
    return None


def get_events(ws_events) -> List[Tuple[str, str]]:
    hdr = headers_of(ws_events)
    name_idx = find_col(hdr, ["Event", "Card", "EventName", "Name", "Card Name", "Event Name"])  # display
    id_idx = find_col(hdr, ["EventId", "Event ID", "EventID", "Event Id", "eventId"])  # id
    if name_idx is None and id_idx is None:
        raise SystemExit("Events sheet must have Event or EventId column")
    seen = set()
    result: List[Tuple[str, str]] = []
    for row in ws_events.iter_rows(min_row=2, values_only=True):
        name = str(row[name_idx - 1]).strip() if name_idx and row[name_idx - 1] is not None else ""
        eid = str(row[id_idx - 1]).strip() if id_idx and row[id_idx - 1] is not None else ""
        if not name and not eid:
            continue
        display = name or eid
        if display in seen:
            continue
        seen.add(display)
        result.append((display, eid))
        if len(result) >= 5000:
            break
    return result


def get_fighters_by_event(ws_profiles, events: List[Tuple[str, str]]) -> Dict[str, List[str]]:
    hdr = headers_of(ws_profiles)
    ev_name_idx = find_col(hdr, ["Event", "Card", "EventName", "Name", "Card Name", "Event Name"])  # optional
    ev_id_idx = find_col(hdr, ["EventId", "Event ID", "EventID", "Event Id", "eventId"])
    fighter_idx = find_col(hdr, ["Profile", "profile", "Name", "Fighter"])
    if fighter_idx is None:
        raise SystemExit("Profiles sheet must have a fighter name column (Profile/Name/Fighter)")

    fighters: Dict[str, List[str]] = {name: [] for name, _ in events}

    # Build fast lookup for event name -> id (if present)
    name_to_id: Dict[str, str] = {name: eid for name, eid in events}

    for row in ws_profiles.iter_rows(min_row=2, values_only=True):
        fighter = str(row[fighter_idx - 1]).strip() if row[fighter_idx - 1] is not None else ""
        if not fighter:
            continue
        prof_ev_name = str(row[ev_name_idx - 1]).strip() if ev_name_idx and row[ev_name_idx - 1] is not None else ""
        prof_ev_id = str(row[ev_id_idx - 1]).strip() if ev_id_idx and row[ev_id_idx - 1] is not None else ""

        target_event_name: Optional[str] = None
        if prof_ev_name and prof_ev_name in fighters:
            target_event_name = prof_ev_name
        else:
            # match by id -> name
            for name, eid in events:
                if eid and eid == prof_ev_id:
                    target_event_name = name
                    break
        if not target_event_name:
            continue
        bucket = fighters[target_event_name]
        if fighter not in bucket:
            bucket.append(fighter)
    return fighters


def main() -> None:
    ap = argparse.ArgumentParser(description="Create cascading dropdowns: Event (B2) -> Fighter (C2)")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--events-sheet", default="Events")
    ap.add_argument("--profiles-sheet", default="Profiles")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, keep_vba=True, data_only=False)
    if args.events_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.events_sheet}")
    if args.profiles_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.profiles_sheet}")
    ws_events = wb[args.events_sheet]
    ws_profiles = wb[args.profiles_sheet]

    ws_lists = wb["Lists"] if "Lists" in wb.sheetnames else wb.create_sheet("Lists")
    if "Dashboard" in wb.sheetnames:
        ws_dash = wb["Dashboard"]
    else:
        ws_dash = wb.create_sheet("Dashboard")

    # Gather data
    events = get_events(ws_events)
    fighters_by_event = get_fighters_by_event(ws_profiles, events)

    # Clear Lists
    ws_lists.delete_rows(1, ws_lists.max_row or 1)

    # Write Event list and mapping
    ws_lists["A1"] = "EventNames"
    ws_lists["B1"] = "EventName"
    ws_lists["C1"] = "EventKey"
    for i, (ev_name, _) in enumerate(events, start=2):
        ws_lists.cell(row=i, column=1, value=ev_name)
        ws_lists.cell(row=i, column=2, value=ev_name)
        ws_lists.cell(row=i, column=3, value=sanitize_key(ev_name))
    last_row = 1 + len(events)

    # Write per-event fighter columns starting at column E
    col = 5  # E
    from openpyxl.utils import get_column_letter  # type: ignore
    from openpyxl.workbook.defined_name import DefinedName  # type: ignore

    # Remove prior Fighters_* names if present
    try:
        to_remove = []
        for n in wb.defined_names.definedName:
            if n.name and n.name.startswith("Fighters_"):
                to_remove.append(n)
        for n in to_remove:
            wb.defined_names.definedName.remove(n)
    except Exception:
        pass

    for ev_name, _ in events:
        fighters = fighters_by_event.get(ev_name, [])
        col_letter = get_column_letter(col)
        header_cell = f"{col_letter}1"
        ws_lists[header_cell] = f"Fighters_{sanitize_key(ev_name)}"
        # Write fighters under header
        for i, v in enumerate(fighters, start=2):
            ws_lists.cell(row=i, column=col, value=v)
        last_f_row = 1 + len(fighters)
        if last_f_row < 2:
            last_f_row = 2
        # Create named range for this event's fighter list
        ref = f"'Lists'!${col_letter}$2:${col_letter}${last_f_row}"
        name = f"Fighters_{sanitize_key(ev_name)}"
        try:
            wb.defined_names.add(DefinedName(name=name, attr_text=ref))
        except Exception:
            # Fallback append
            wb.defined_names.append(DefinedName(name=name, attr_text=ref))
        col += 1

    # Event dropdown in Dashboard!B2
    ws_dash["B1"] = "Event"
    dv_event = DataValidation(type="list", formula1=f"='Lists'!$A$2:$A${last_row}", allow_blank=True)
    ws_dash.add_data_validation(dv_event)
    dv_event.add(ws_dash["B2"])

    # Fighter dropdown in Dashboard!C2 using INDIRECT to named range by sanitized mapping
    ws_dash["C1"] = "Fighter"
    # Formula: =INDIRECT("Fighters_" & INDEX(Lists!$C$2:$C$N, MATCH(B2, Lists!$B$2:$B$N, 0)))
    dv_fighter = DataValidation(
        type="list",
        formula1=f"=INDIRECT(\"Fighters_\" & INDEX(Lists!$C$2:$C${last_row}, MATCH(B2, Lists!$B$2:$B${last_row}, 0)))",
        allow_blank=True,
    )
    ws_dash.add_data_validation(dv_fighter)
    dv_fighter.add(ws_dash["C2"])

    ws_dash.freeze_panes = "A2"

    wb.save(args.workbook)
    print(f"Wrote {len(events)} events and cascading fighter dropdown. B2 -> C2 ready.")


if __name__ == "__main__":
    main()


