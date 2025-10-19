import argparse
from typing import List, Optional

from openpyxl import load_workbook  # type: ignore
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore


def find_col_index(headers: List[str], candidates: List[str]) -> Optional[int]:
    lookup = { (h or "").strip().lower(): i for i, h in enumerate(headers, start=1) }
    for name in candidates:
        idx = lookup.get(name.strip().lower())
        if idx is not None:
            return idx
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a single Event dropdown at Dashboard!B2")
    parser.add_argument("--workbook", required=True, help="Path to workbook (.xlsm/.xlsx)")
    parser.add_argument("--events-sheet", default="Events", help="Events sheet name")
    parser.add_argument("--target", default="Dashboard!B2", help="Target like Sheet!Cell (default Dashboard!B2)")
    args = parser.parse_args()

    wb = load_workbook(args.workbook, keep_vba=True, data_only=False)

    if args.events_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.events_sheet}")
    ws_events = wb[args.events_sheet]

    # Ensure helper sheets exist
    ws_lists = wb["Lists"] if "Lists" in wb.sheetnames else wb.create_sheet("Lists")

    # Parse target
    if "!" in args.target:
        target_sheet, target_cell = args.target.split("!", 1)
    else:
        target_sheet, target_cell = "Dashboard", args.target
    ws_target = wb[target_sheet] if target_sheet in wb.sheetnames else wb.create_sheet(target_sheet)

    # Detect Event or EventId column
    headers: List[str] = [str(c.value).strip() if c.value is not None else "" for c in ws_events[1]]
    col_idx = find_col_index(headers, [
        "Event", "Card", "EventName", "Name", "Card Name", "Event Name",
    ])
    if col_idx is None:
        col_idx = find_col_index(headers, [
            "EventId", "Event ID", "EventID", "Event Id", "eventId",
        ])
    if col_idx is None:
        raise SystemExit("Could not find Event or EventId column in Events sheet")

    # Gather unique values, up to 5000, skipping blanks
    seen = set()
    values: List[str] = []
    for r in ws_events.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
        v = r[0]
        s = (str(v).strip() if v is not None else "")
        if not s:
            continue
        if s in seen:
            continue
        seen.add(s)
        values.append(s)
        if len(values) >= 5000:
            break

    # Write list to Lists!A
    ws_lists.delete_rows(1, ws_lists.max_row or 1)
    ws_lists["A1"] = "EventNames"
    for i, v in enumerate(values, start=2):
        ws_lists.cell(row=i, column=1, value=v)

    # Apply data validation to target cell only
    dv = DataValidation(
        type="list",
        formula1=f"='Lists'!$A$2:$A${len(values)+1}",
        allow_blank=True,
    )
    ws_target.add_data_validation(dv)
    dv.add(ws_target[target_cell])

    # Minimal headers on Dashboard
    if target_sheet == "Dashboard":
        ws_target["B1"] = "Event"
        ws_target.freeze_panes = "A2"

    wb.save(args.workbook)
    print(f"Dropdown written to {args.target} with {len(values)} options.")


if __name__ == "__main__":
    main()


