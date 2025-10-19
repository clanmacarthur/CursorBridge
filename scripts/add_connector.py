import argparse
from typing import Dict, List, Optional, Tuple
import difflib

from openpyxl import load_workbook  # type: ignore


ROLES = [
    ("event_key", "Events", ["EventId", "Event ID", "EventID", "Event Id", "eventId"]),
    (
        "event_display",
        "Events",
        [
            "Event",
            "Card",
            "EventName",
            "Name",
            "Card Name",
            "Event Name",
            "Column3",
            "Column 3",
            "Card Events",
            "Card events",
            "Card Event",
            "Card event",
        ],
    ),
    ("fighter_name", "Profiles", ["Profile", "profile", "Name", "Fighter"]),
]

# Stats we bind
STAT_KEYS = [
    "Height", "Weight", "Reach", "Stance", "Record",
    "TSSL", "TSSA", "STRACC%", "SSL/M", "SSA/M", "DSL/M", "DSA/M", "KD/15mins",
]


def headers_of(ws) -> List[str]:
    return [str(c.value).strip() if c.value is not None else "" for c in ws[1]]


def detect_header(headers: List[str], candidates: List[str]) -> Tuple[str, int]:
    lower = {(h or "").strip().lower(): (h, i) for i, h in enumerate(headers, start=1)}
    for cand in candidates:
        r = lower.get(cand.strip().lower())
        if r:
            return r[0], r[1]
    return "", 0


def fuzzy_match_header(headers: List[str], target: str, threshold: float = 0.77) -> Tuple[str, int, float]:
    """Return (found_header, col_idx, score) using fuzzy similarity if >= threshold, else ("", 0, 0)."""
    best_score = 0.0
    best_name = ""
    best_idx = 0
    for i, h in enumerate(headers, start=1):
        if not h:
            continue
        score = difflib.SequenceMatcher(None, h.strip().lower(), target.strip().lower()).ratio()
        if score > best_score:
            best_score = score
            best_name = h
            best_idx = i
    if best_score >= threshold:
        return best_name, best_idx, best_score
    return "", 0, 0.0


def main() -> None:
    ap = argparse.ArgumentParser(description="Create Connector sheet showing column mappings")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--events-sheet", default="Events")
    ap.add_argument("--profiles-sheet", default="Profiles")
    ap.add_argument("--stats-sheet", default=None, help="Optional stats sheet (e.g., FightsStats)")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, keep_vba=True, data_only=False)

    if args.events_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.events_sheet}")
    if args.profiles_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.profiles_sheet}")

    ws_events = wb[args.events_sheet]
    ws_profiles = wb[args.profiles_sheet]
    stats_sheet_name = args.stats_sheet or args.profiles_sheet
    if stats_sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {stats_sheet_name}")
    ws_stats = wb[stats_sheet_name]

    ev_hdr = headers_of(ws_events)
    pf_hdr = headers_of(ws_profiles)
    st_hdr = headers_of(ws_stats)

    rows: List[List[str]] = []
    # Core roles
    for role, sheet, candidates in ROLES:
        hdr_list = ev_hdr if sheet == args.events_sheet else pf_hdr
        found_name, col_idx = detect_header(hdr_list, candidates)
        note = "EXACT" if col_idx else ""
        if not col_idx:
            # Try fuzzy on the first candidate term as a representative
            if candidates:
                f_name, f_idx, score = fuzzy_match_header(hdr_list, candidates[0])
                if f_idx:
                    found_name, col_idx = f_name, f_idx
                    note = f"FUZZY:{score:.2f}"
        rows.append([role, sheet, ", ".join(candidates), found_name or "", str(col_idx or ""), "OK" if col_idx else "MISSING", note])

    # Stats
    for key in STAT_KEYS:
        # Prefer stats sheet, fall back to profiles; then fuzzy
        used_sheet = stats_sheet_name
        found_name, col_idx = detect_header(st_hdr, [key])
        note = "EXACT" if col_idx else ""
        if not col_idx:
            f_name, f_idx, score = fuzzy_match_header(st_hdr, key)
            if f_idx:
                found_name, col_idx = f_name, f_idx
                note = f"FUZZY:{score:.2f}"
        if not col_idx:
            used_sheet = args.profiles_sheet
            found_name, col_idx = detect_header(pf_hdr, [key])
            note = "EXACT" if col_idx else note
            if not col_idx:
                f_name, f_idx, score = fuzzy_match_header(pf_hdr, key)
                if f_idx:
                    found_name, col_idx = f_name, f_idx
                    note = f"FUZZY:{score:.2f}"
        rows.append([f"stat:{key}", used_sheet, key, found_name or "", str(col_idx or ""), "OK" if col_idx else "MISSING", note])

    # Create/replace Connector sheet
    if "Connector" in wb.sheetnames:
        wb.remove(wb["Connector"])
    ws = wb.create_sheet("Connector")
    ws["A1"], ws["B1"], ws["C1"], ws["D1"], ws["E1"], ws["F1"], ws["G1"] = (
        "role", "sheet", "expected", "found", "col", "status", "note"
    )
    for i, r in enumerate(rows, start=2):
        for j, v in enumerate(r, start=1):
            ws.cell(row=i, column=j, value=v)

    wb.save(args.workbook)
    print(f"Connector sheet created with {len(rows)} rows.")


if __name__ == "__main__":
    main()


