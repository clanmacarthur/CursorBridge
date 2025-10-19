import argparse
from openpyxl import load_workbook  # type: ignore
from openpyxl.chart import PieChart, Reference  # type: ignore


OFFENSE_KEYS = [
    ("SSL/M", "SSL/M"),
    ("SSA/M", "SSA/M"),
    ("KD/15mins", "KD/15mins"),
]

DEFENSE_KEYS = [
    ("DSL/M", "DSL/M"),
    ("DSA/M", "DSA/M"),
]


def main() -> None:
    ap = argparse.ArgumentParser(description="Add striking pie charts (combined/offense/defense) to Dashboard")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--profiles-sheet", default="Profiles")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, keep_vba=True, data_only=False)
    if "Dashboard" not in wb.sheetnames:
        raise SystemExit("Dashboard sheet not found. Run bind_stats first.")
    ws = wb["Dashboard"]

    # Build helper ranges for chart labels/values in an out-of-the-way area (e.g., AA columns)
    start_col = 27  # AA
    r = 2
    ws.cell(row=1, column=start_col, value="Chart_Label")
    ws.cell(row=1, column=start_col + 1, value="Chart_Value")

    # Combined (SSL/M, SSA/M, DSL/M, DSA/M)
    combined = [
        ("SSL/M", "SSL/M"),
        ("SSA/M", "SSA/M"),
        ("DSL/M", "DSL/M"),
        ("DSA/M", "DSA/M"),
    ]
    for label, key in combined:
        ws.cell(row=r, column=start_col, value=label)
        # Find the label row in the left table (col D labels, col E values)
        ws.cell(row=r, column=start_col + 1, value=f"=IFERROR(INDEX($E:$E, MATCH(\"{key}\", $D:$D, 0)), \"\")")
        r += 1

    # Offense group
    offense_row_start = r + 1
    for label, key in OFFENSE_KEYS:
        ws.cell(row=r, column=start_col, value=label)
        ws.cell(row=r, column=start_col + 1, value=f"=IFERROR(INDEX($E:$E, MATCH(\"{key}\", $D:$D, 0)), \"\")")
        r += 1
    offense_row_end = r - 1

    # Defense group
    defense_row_start = r + 1
    for label, key in DEFENSE_KEYS:
        ws.cell(row=r, column=start_col, value=label)
        ws.cell(row=r, column=start_col + 1, value=f"=IFERROR(INDEX($E:$E, MATCH(\"{key}\", $D:$D, 0)), \"\")")
        r += 1
    defense_row_end = r - 1

    def add_pie(top_left_col: int, top_left_row: int, title: str, label_start_row: int, label_end_row: int):
        labels = Reference(ws, min_col=start_col, min_row=label_start_row, max_row=label_end_row)
        values = Reference(ws, min_col=start_col + 1, min_row=label_start_row, max_row=label_end_row)
        chart = PieChart()
        chart.title = title
        chart.add_data(values, titles_from_data=False)
        chart.set_categories(labels)
        # Position chart roughly at cell
        ws.add_chart(chart, ws.cell(row=top_left_row, column=top_left_col).coordinate)

    # Place charts on the dashboard
    # Combined
    add_pie(8, 4, "Striking (Combined)", 2, 5)  # around H4
    # Offense
    add_pie(8, 20, "Striking Offense", offense_row_start, offense_row_end)  # H20
    # Defense
    add_pie(15, 20, "Striking Defense", defense_row_start, defense_row_end)  # O20

    wb.save(args.workbook)
    print("Striking pie charts added.")


if __name__ == "__main__":
    main()


