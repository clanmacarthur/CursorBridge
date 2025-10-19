import argparse
from typing import Dict, List, Optional

from openpyxl import load_workbook  # type: ignore


MEASURABLES = [
    ("Height", "Height"),
    ("Weight", "Weight"),
    ("Reach", "Reach"),
    ("Stance", "Stance"),
    ("Record", "Record"),
]

STRIKING = [
    ("TSSL", "TSSL"),
    ("TSSA", "TSSA"),
    ("STRACC%", "STRACC%"),
    ("SSL/M", "SSL/M"),
    ("SSA/M", "SSA/M"),
    ("DSL/M", "DSL/M"),
    ("DSA/M", "DSA/M"),
    ("KD/15mins", "KD/15mins"),
]


def headers_of(ws) -> List[str]:
    return [str(c.value).strip() if c.value is not None else "" for c in ws[1]]


def find_col(headers: List[str], key: str) -> Optional[int]:
    k = key.strip().lower()
    for i, h in enumerate(headers, start=1):
        if (h or "").strip().lower() == k:
            return i
    return None


def main() -> None:
    ap = argparse.ArgumentParser(description="Bind measurables and striking stats to Dashboard based on C2 selection")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--profiles-sheet", default="Profiles")
    ap.add_argument("--stats-sheet", default=None, help="Optional sheet to source stats (e.g., FightsStats)")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, keep_vba=True, data_only=False)

    if args.profiles_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.profiles_sheet}")
    ws_profiles = wb[args.profiles_sheet]

    ws_dash = wb["Dashboard"] if "Dashboard" in wb.sheetnames else wb.create_sheet("Dashboard")

    # Choose stats source sheet (defaults to profiles)
    stats_sheet_name = args.stats_sheet or args.profiles_sheet
    if stats_sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {stats_sheet_name}")
    ws_stats = wb[stats_sheet_name]

    # Find fighter name column on stats sheet; if missing, fall back to profiles (but require name column)
    hdr_stats = headers_of(ws_stats)
    fighter_idx_stats: Optional[int] = None
    for key in ["Profile", "profile", "Name", "Fighter"]:
        fighter_idx_stats = find_col(hdr_stats, key)
        if fighter_idx_stats:
            break

    hdr_profiles = headers_of(ws_profiles)
    fighter_idx_profiles: Optional[int] = None
    for key in ["Profile", "profile", "Name", "Fighter"]:
        fighter_idx_profiles = find_col(hdr_profiles, key)
        if fighter_idx_profiles:
            break
    if not fighter_idx_stats and not fighter_idx_profiles:
        raise SystemExit("Neither stats sheet nor profiles sheet has a Profile/Name/Fighter column")

    # Build header maps for both sheets
    need = {k for _, k in MEASURABLES + STRIKING}
    stats_colmap: Dict[str, int] = {}
    for name in need:
        idx = find_col(hdr_stats, name)
        if idx:
            stats_colmap[name] = idx
    profiles_colmap: Dict[str, int] = {}
    for name in need:
        idx = find_col(hdr_profiles, name)
        if idx:
            profiles_colmap[name] = idx

    # Optional: Connector overrides (lets you remap columns without code changes)
    connector_overrides: Dict[str, Dict[str, int]] = {}
    connector_sheets_for_stat: Dict[str, str] = {}
    if "Connector" in wb.sheetnames:
        ws_conn = wb["Connector"]
        # Build a cache of headers per sheet for quick lookup
        headers_cache: Dict[str, List[str]] = {}
        def get_headers(sheet_name: str) -> List[str]:
            if sheet_name not in headers_cache:
                headers_cache[sheet_name] = headers_of(wb[sheet_name])
            return headers_cache[sheet_name]

        for row in ws_conn.iter_rows(min_row=2, values_only=True):
            role = str(row[0]).strip() if row[0] is not None else ""
            sheet_name = str(row[1]).strip() if row[1] is not None else ""
            expected = str(row[2]).strip() if row[2] is not None else ""
            found = str(row[3]).strip() if row[3] is not None else ""
            col_text = str(row[4]).strip() if row[4] is not None else ""
            status = str(row[5]).strip().upper() if row[5] is not None else ""
            if not role.startswith("stat:"):
                continue
            key = role.split(":", 1)[1]
            if status != "OK" or not sheet_name:
                continue
            # Determine column index
            col_index: Optional[int] = None
            try:
                if col_text:
                    col_index = int(col_text)
            except Exception:
                col_index = None
            if not col_index:
                # try by header name
                hdr = get_headers(sheet_name)
                col_index = find_col(hdr, found) if found else None
            if not col_index:
                continue
            if sheet_name not in connector_overrides:
                connector_overrides[sheet_name] = {}
            connector_overrides[sheet_name][key] = col_index
            connector_sheets_for_stat[key] = sheet_name

    # Helper to get fighter name column for a given sheet
    def get_name_col_for_sheet(sheet_name: str) -> Optional[int]:
        hdr = headers_of(wb[sheet_name])
        for key in ["Profile", "profile", "Name", "Fighter"]:
            idx = find_col(hdr, key)
            if idx:
                return idx
        return None

    # Layout sections on Dashboard
    ws_dash["E1"] = "Selected Fighter"
    ws_dash["E2"] = "=" + "C2"  # show selected fighter name

    def make_index_formula(sheet_name: str, value_col: int, name_col: int) -> str:
        from_sheet = wb[sheet_name]
        val_letter = from_sheet.cell(row=1, column=value_col).column_letter
        name_letter = from_sheet.cell(row=1, column=name_col).column_letter
        return (
            f"=IFERROR(INDEX('{sheet_name}'!${val_letter}:${val_letter}, "
            f"MATCH($C$2, '{sheet_name}'!${name_letter}:${name_letter}, 0)), \"\")"
        )

    start_row = 4
    ws_dash["D3"] = "MEASURABLES"
    r = start_row
    for label, key in MEASURABLES:
        ws_dash.cell(row=r, column=4, value=label)
        # Connector override per-stat takes priority
        if key in connector_sheets_for_stat:
            sheet_for_key = connector_sheets_for_stat[key]
            name_idx = get_name_col_for_sheet(sheet_for_key)
            col_idx = connector_overrides.get(sheet_for_key, {}).get(key)
            if name_idx and col_idx:
                ws_dash.cell(row=r, column=5).value = make_index_formula(sheet_for_key, col_idx, name_idx)
        elif key in stats_colmap and fighter_idx_stats:
            ws_dash.cell(row=r, column=5).value = make_index_formula(stats_sheet_name, stats_colmap[key], fighter_idx_stats)
        elif key in profiles_colmap and fighter_idx_profiles:
            ws_dash.cell(row=r, column=5).value = make_index_formula(args.profiles_sheet, profiles_colmap[key], fighter_idx_profiles)
        r += 1

    r += 1
    ws_dash.cell(row=r-1, column=4, value="STRIKING")
    for label, key in STRIKING:
        ws_dash.cell(row=r, column=4, value=label)
        if key in connector_sheets_for_stat:
            sheet_for_key = connector_sheets_for_stat[key]
            name_idx = get_name_col_for_sheet(sheet_for_key)
            col_idx = connector_overrides.get(sheet_for_key, {}).get(key)
            if name_idx and col_idx:
                ws_dash.cell(row=r, column=5).value = make_index_formula(sheet_for_key, col_idx, name_idx)
        elif key in stats_colmap and fighter_idx_stats:
            ws_dash.cell(row=r, column=5).value = make_index_formula(stats_sheet_name, stats_colmap[key], fighter_idx_stats)
        elif key in profiles_colmap and fighter_idx_profiles:
            ws_dash.cell(row=r, column=5).value = make_index_formula(args.profiles_sheet, profiles_colmap[key], fighter_idx_profiles)
        r += 1

    wb.save(args.workbook)
    print("Bound measurables and striking stats to Dashboard.")


if __name__ == "__main__":
    main()


