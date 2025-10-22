import importlib
import os
import time
import re
from typing import Optional
from typing import Any, Callable


def _open_book(path: str) -> Any:
    # Lazy import to avoid requiring xlwings for non-Excel commands
    import xlwings as xw  # type: ignore
    visible_env = os.getenv("XLWINGS_VISIBLE", "false").lower() == "true"
    app = xw.App(visible=visible_env, add_book=False)
    # Lower macro security for automation before opening the workbook (1 = msoAutomationSecurityLow)
    try:
        app.api.AutomationSecurity = 1
        app.api.DisplayAlerts = False
    except Exception:
        # Best effort; continue even if not available
        pass
    try:
        book = app.books.open(path)
    except Exception:
        app.quit()
        raise
    return book


def run_macro(workbook_path: str, macro_name: str) -> None:
    book = _open_book(workbook_path)
    try:
        app = book.app
        wb_name = os.path.basename(workbook_path)
        wb_path = os.path.abspath(workbook_path)
        # Candidate macro name variants
        candidates = []
        if macro_name:
            candidates.append(macro_name)
            # Also try bare procedure name (without module qualifier)
            bare = macro_name.split(".")[-1]
            if bare and bare not in candidates:
                candidates.append(bare)
            if "!" not in macro_name:
                candidates.append(f"{wb_name}!{macro_name}")
                candidates.append(f"'{wb_name}'!{macro_name}")
                # Full path variants
                candidates.append(f"{wb_path}!{macro_name}")
                candidates.append(f"'{wb_path}'!{macro_name}")

        last_err: Exception | None = None
        for cand in candidates:
            try:
                if "!" in cand:
                    # Use Application.Run for workbook-qualified names
                    app.api.Run(cand)
                else:
                    book.macro(cand)()
                book.save()
                break
            except Exception as ex:  # try next candidate
                last_err = ex
        else:
            if last_err is not None:
                raise last_err
    finally:
        book.close()
        book.app.quit()


def run_python(workbook_path: str, callable_path: str) -> None:
    try:
        module_name, func_name = callable_path.split(":", 1)
    except ValueError:
        raise ValueError("callable path must be in the form 'module:function'")

    # Try the xlwings path first
    try:
        module = importlib.import_module(module_name)
        func: Callable[[Any], None] = getattr(module, func_name)
        book = _open_book(workbook_path)
        try:
            func(book)
            book.save()
        finally:
            book.close()
            book.app.quit()
        return
    except ModuleNotFoundError as e:
        if e.name != "xlwings":
            # Module of the callable is missing, bubble up
            raise
        # Fall through to openpyxl fallback

    # Fallbacks using openpyxl while keeping macros
    if module_name == "cb.excel" and func_name == "example_task":
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as ex:
            raise RuntimeError("openpyxl not installed; run: python -m pip install openpyxl") from ex
        wb = load_workbook(workbook_path, keep_vba=True)
        ws = wb.worksheets[0]
        ws["A1"] = f"Updated by CursorBridge at {time.strftime('%Y-%m-%d %H:%M:%S')}"
        wb.save(workbook_path)
        return

    if module_name == "cb.excel" and func_name == "clean_dates_col5":
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as ex:
            raise RuntimeError("openpyxl not installed; run: python -m pip install openpyxl") from ex
        wb = load_workbook(workbook_path, keep_vba=True)
        col_index_env = os.getenv("CB_COL", "5")
        try:
            col_index = int(col_index_env)
            if col_index <= 0:
                raise ValueError
        except ValueError:
            raise ValueError(f"Invalid CB_COL value: {col_index_env}")
        sheet_filter = os.getenv("CB_SHEET")
        pattern = re.compile(r"(\b\d{1,2})(st|nd|rd|th)(,?)", re.IGNORECASE)

        target_sheets = []
        if sheet_filter:
            if sheet_filter not in wb.sheetnames:
                raise KeyError(f"Worksheet '{sheet_filter}' not found")
            target_sheets = [wb[sheet_filter]]
        else:
            target_sheets = wb.worksheets

        total_changed = 0
        for ws in target_sheets:
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_index)
                cell_val = cell.value
                if isinstance(cell_val, str) and cell_val:
                    new_val = pattern.sub(r"\1,", cell_val)
                    if new_val != cell_val:
                        cell.value = new_val
                        total_changed += 1
        wb.save(workbook_path)
        msg = f"Cleaned {total_changed} cell(s) in column {col_index}"
        if sheet_filter:
            msg += f" on sheet '{sheet_filter}'"
        print(msg)
        return

    raise RuntimeError(
        "xlwings is not available and no fallback is defined for this callable. "
        "Install xlwings: python -m pip install xlwings"
    )


def example_task(book: Any) -> None:
    sheet = book.sheets[0]
    sheet.range("A1").value = f"Updated by CursorBridge at {time.strftime('%Y-%m-%d %H:%M:%S')}"


def _parse_vba_module_name(module_path: str) -> str:
    try:
        with open(module_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                m = re.match(r"\s*Attribute\s+VB_Name\s*=\s*\"([^\"]+)\"", line)
                if m:
                    return m.group(1)
    except FileNotFoundError:
        raise FileNotFoundError(f"VBA module not found: {module_path}")
    # Fallback to filename stem
    return os.path.splitext(os.path.basename(module_path))[0]


def vba_import_module(workbook_path: str, module_path: str, module_name: Optional[str] = None, replace: bool = True) -> None:
    """Import a VBA module (.bas/.cls) into a workbook via COM (Excel hidden).

    Requires Excel setting: Trust access to the VBA project object model.
    """
    try:
        import xlwings as xw  # type: ignore
    except Exception as ex:
        raise RuntimeError("xlwings not installed; run: python -m pip install xlwings") from ex

    if not os.path.exists(workbook_path):
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if not os.path.exists(module_path):
        raise FileNotFoundError(f"Module file not found: {module_path}")

    mod_name = module_name or _parse_vba_module_name(module_path)

    app = xw.App(visible=False, add_book=False)
    try:
        book = app.books.open(workbook_path)
        try:
            vbproj = book.api.VBProject
        except Exception as ex:
            raise RuntimeError(
                "Access to VBProject denied. In Excel, enable: Trust Center > Trust Center Settings > Macro Settings > 'Trust access to the VBA project object model'."
            ) from ex

        components = vbproj.VBComponents
        if replace:
            try:
                comp = components(mod_name)
                components.Remove(comp)
            except Exception:
                pass  # not present, fine
        try:
            components.Import(module_path)
        except Exception as ex:
            raise RuntimeError(f"Failed to import VBA module '{mod_name}' from {module_path}") from ex
        book.save()
    finally:
        try:
            book.close()
        finally:
            app.quit()

def clean_dates_col5(book: Any) -> None:
    """Clean ordinal date suffixes (st/nd/rd/th) in column 5 for PAD.

    - Targets column 5 by default (override with env CB_COL)
    - Targets all sheets by default (restrict with env CB_SHEET)
    - Replaces patterns like '1st', '2nd', '3rd', '4th' → '1,', '2,', ...
    """
    col_index_env = os.getenv("CB_COL", "5")
    try:
        col_index = int(col_index_env)
        if col_index <= 0:
            raise ValueError
    except ValueError:
        raise ValueError(f"Invalid CB_COL value: {col_index_env}")

    sheet_filter = os.getenv("CB_SHEET")
    # Replace ordinal suffix and any immediate comma with a single comma
    pattern = re.compile(r"(\b\d{1,2})(st|nd|rd|th)(,?)", re.IGNORECASE)

    target_sheets = [s for s in book.sheets if (sheet_filter is None or s.name == sheet_filter)]
    total_changed = 0
    for sheet in target_sheets:
        used = sheet.used_range
        # If the sheet is empty, skip
        if used is None:
            continue
        last_row = used.last_cell.row
        if last_row < 1:
            continue

        rng = sheet.range((1, col_index), (last_row, col_index))
        values = rng.options(ndim=2).value or []
        if not values:
            continue

        changed_any = False
        for i in range(len(values)):
            cell_val = values[i][0]
            if isinstance(cell_val, str) and cell_val:
                new_val = pattern.sub(r"\1,", cell_val)
                if new_val != cell_val:
                    values[i][0] = new_val
                    changed_any = True
                    total_changed += 1
        if changed_any:
            rng.value = values

    msg = f"Cleaned {total_changed} cell(s) in column {col_index}"
    if sheet_filter:
        msg += f" on sheet '{sheet_filter}'"
    print(msg)

def _openpyxl_get_worksheet(workbook, sheet: str):
    # sheet can be a name or 1-based index as string
    if sheet.isdigit():
        idx = int(sheet)
        if idx <= 0:
            raise ValueError("Sheet index must be >= 1")
        return workbook.worksheets[idx - 1]
    if sheet in workbook.sheetnames:
        return workbook[sheet]
    raise KeyError(f"Worksheet '{sheet}' not found")


def write_cell_openpyxl(workbook_path: str, sheet: str, cell: str, value: str) -> None:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as ex:
        raise RuntimeError("openpyxl not installed; run: python -m pip install openpyxl") from ex
    wb = load_workbook(workbook_path, keep_vba=True)
    ws = _openpyxl_get_worksheet(wb, sheet)
    ws[cell] = value
    wb.save(workbook_path)


def append_row_openpyxl(workbook_path: str, sheet: str, values: list[str]) -> None:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as ex:
        raise RuntimeError("openpyxl not installed; run: python -m pip install openpyxl") from ex
    wb = load_workbook(workbook_path, keep_vba=True)
    ws = _openpyxl_get_worksheet(wb, sheet)
    ws.append(values)
    wb.save(workbook_path)


