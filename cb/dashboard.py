import os
import shlex
import subprocess
from typing import List, Optional


def _ps_escape(path: str) -> str:
    return path.replace("'", "''")


def _build_dashboard_safe(
    workbook_path: str,
    *,
    events_sheet: str,
    fighters_sheet: str,
    fights_sheet: str,
) -> None:
    """Safe mode: create a minimal dashboard with a static Event dropdown.

    - No COM, no PowerShell, no dynamic array formulas
    - Uses openpyxl only; preserves macros via keep_vba=True
    """
    try:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    except Exception as ex:  # pragma: no cover
        raise RuntimeError("openpyxl not installed; run: python -m pip install openpyxl") from ex

    def find_col_index_case_insensitive(headers: List[str], candidates: List[str]) -> Optional[int]:
        lowered = { (h or "").strip().lower(): i for i, h in enumerate(headers) }
        for cand in candidates:
            idx = lowered.get(cand.strip().lower())
            if idx is not None:
                return idx
        return None

    wb = load_workbook(workbook_path, keep_vba=True, data_only=False)
    # Get or create Events sheet
    if events_sheet in wb.sheetnames:
        ws_events = wb[events_sheet]
    else:
        raise FileNotFoundError(f"Sheet not found: {events_sheet}")

    # Prepare Dashboard and Lists sheets (recreate to avoid lingering validations)
    if 'Dashboard' in wb.sheetnames:
        wb.remove(wb['Dashboard'])
    ws_dash = wb.create_sheet('Dashboard')
    if 'Lists' in wb.sheetnames:
        wb.remove(wb['Lists'])
    ws_lists = wb.create_sheet('Lists')

    # Read headers in Events (first row)
    headers: List[str] = []
    for cell in ws_events[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    event_name_idx = find_col_index_case_insensitive(headers, [
        'Event','Card','EventName','Name','CardName','Card Name','Event Name'
    ])
    if event_name_idx is None:
        # fallback to EventId so at least a dropdown appears
        event_name_idx = find_col_index_case_insensitive(headers, [
            'EventId','Event ID','EventID','Event Id','eventId'
        ])
    if event_name_idx is None:
        raise RuntimeError("Could not find an Event or EventId column in Events sheet")

    # Collect unique, non-empty values from that column; stop after long blank streak for speed
    seen = set()
    unique_vals: List[str] = []
    consecutive_blanks = 0
    max_rows_scan = min(ws_events.max_row, 5000)
    for r in range(2, max_rows_scan + 1):
        cell = ws_events.cell(row=r, column=event_name_idx + 1)
        val = cell.value
        s = str(val).strip() if val is not None else ""
        if not s:
            consecutive_blanks += 1
            if consecutive_blanks > 500:
                break
            continue
        consecutive_blanks = 0
        if s not in seen:
            seen.add(s)
            unique_vals.append(s)
        if len(unique_vals) >= 2000:
            break

    # Write Lists!A1 header and values starting A2
    ws_lists['A1'] = 'EventNames'
    for i, v in enumerate(unique_vals, start=2):
        ws_lists.cell(row=i, column=1, value=v)

    # Dashboard headers and data validation
    ws_dash['B1'] = 'Event'
    dv = DataValidation(type='list', formula1=f"='Lists'!$A$2:$A${len(unique_vals)+1}", allow_blank=True)
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash['B2'])

    # Freeze top row for clarity
    ws_dash.freeze_panes = 'A2'

    wb.save(workbook_path)


def build_dashboard(
    workbook_path: str,
    visible: bool = False,
    *,
    events_sheet: str = "Events",
    fighters_sheet: str = "fighters",
    fights_sheet: str = "fights",
    safe: bool = False,
) -> None:
    if safe:
        _build_dashboard_safe(
            workbook_path,
            events_sheet=events_sheet,
            fighters_sheet=fighters_sheet,
            fights_sheet=fights_sheet,
        )
        return
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(workbook_path)

    p = _ps_escape(workbook_path)
    ev = _ps_escape(events_sheet)
    fi = _ps_escape(fighters_sheet)
    ft = _ps_escape(fights_sheet)
    vis = "$true" if visible else "$false"

    # PowerShell script builds Dashboard and Lists sheets, dynamic dropdowns,
    # and a placeholder pie chart area wired to dynamic lists.
    ps = f"$ErrorActionPreference='Stop';$path='{p}';$vis={vis};$eventsName='{ev}';$fightersName='{fi}';$fightsName='{ft}';"
    ps += r'''
$xl = New-Object -ComObject Excel.Application
$xl.Visible = $vis
$xl.DisplayAlerts = $false
$xl.AskToUpdateLinks = $false
$wb = $xl.Workbooks.Open($path, 0)  # UpdateLinks=0

function Get-ColIndexByHeader($ws, $header) {
  $ur = $ws.UsedRange
  $cols = $ur.Columns.Count
  for ($c = 1; $c -le $cols; $c++) {
    $val = [string]($ws.Cells.Item(1,$c).Value2)
    if (($val.Trim()) -ieq $header) { return $c }
  }
  return $null
}

function Get-ColIndexByAnyHeader($ws, $headers) {
  foreach ($h in $headers) {
    $idx = Get-ColIndexByHeader $ws $h
    if ($idx) { return $idx }
  }
  return $null
}

function ColLetter($n) {
  $s = ""
  while ($n -gt 0) {
    $n = $n - 1
    # Ensure integer before casting to [char] to avoid Double->Char errors
    $s = [string]([char]([int](65 + ([int]($n % 26))))) + $s
    $n = [math]::Floor($n / 26)
  }
  return $s
}

# Ensure required sheets exist (create if missing)
$wsEvents = $null
try { $wsEvents = $wb.Worksheets.Item($eventsName) } catch {}
if (-not $wsEvents) { $wsEvents = $wb.Worksheets.Add(); $wsEvents.Name = $eventsName }

$wsFighters = $null
try { $wsFighters = $wb.Worksheets.Item($fightersName) } catch {}
if (-not $wsFighters) { $wsFighters = $wb.Worksheets.Add(); $wsFighters.Name = $fightersName }

# Optional fights sheet
$wsFights = $null
try { $wsFights = $wb.Worksheets.Item($fightsName) } catch {}

### Ensure headers and minimal sample data if sheets are empty
# Events headers
$eventIdColIdx = Get-ColIndexByAnyHeader $wsEvents @('EventId','EventID','Event Id','Event ID','eventId')
if (-not $eventIdColIdx) { $wsEvents.Cells.Item(1,1).Value2 = 'EventId'; $eventIdColIdx = 1 }
$eventDisplayColIdx = Get-ColIndexByAnyHeader $wsEvents @('Card','Event','EventName','Name','CardName','Card Name','Event Name')
$eventsRowCount = $wsEvents.UsedRange.Rows.Count
if ($eventsRowCount -lt 3 -and -not [string]($wsEvents.Cells.Item(2,1).Value2)) {
  $wsEvents.Cells.Item(2,1).Value2 = 'E001'
  $wsEvents.Cells.Item(3,1).Value2 = 'E002'
}

# fighters headers
$fightersEventIdColIdx = Get-ColIndexByAnyHeader $wsFighters @('EventId','EventID','Event Id','Event ID','eventId')
$fightersNameColIdx = Get-ColIndexByAnyHeader $wsFighters @('Name','Fighter','Profile','profile')
if (-not $fightersEventIdColIdx) { $wsFighters.Cells.Item(1,1).Value2 = 'EventId'; $fightersEventIdColIdx = 1 }
if (-not $fightersNameColIdx) { $wsFighters.Cells.Item(1,2).Value2 = 'Name'; $fightersNameColIdx = 2 }
$fightersRowCount = $wsFighters.UsedRange.Rows.Count
if ($fightersRowCount -lt 3 -and -not [string]($wsFighters.Cells.Item(2,1).Value2)) {
  $wsFighters.Cells.Item(2,$fightersEventIdColIdx).Value2 = 'E001'
  $wsFighters.Cells.Item(2,$fightersNameColIdx).Value2 = 'Fighter A'
  $wsFighters.Cells.Item(3,$fightersEventIdColIdx).Value2 = 'E002'
  $wsFighters.Cells.Item(3,$fightersNameColIdx).Value2 = 'Fighter B'
}

$fightsEventIdColIdx = $null; $fightsNameColIdx = $null
if ($wsFights) {
  $fightsEventIdColIdx = Get-ColIndexByAnyHeader $wsFights @('EventId','EventID','Event Id','Event ID','eventId')
  $fightsNameColIdx = Get-ColIndexByAnyHeader $wsFights @('FightName','Bout','Fight','Matchup')
  if (-not $fightsEventIdColIdx) { $wsFights.Cells.Item(1,1).Value2 = 'EventId'; $fightsEventIdColIdx = 1 }
  if (-not $fightsNameColIdx) { $wsFights.Cells.Item(1,2).Value2 = 'FightName'; $fightsNameColIdx = 2 }
  $fightsRowCount = $wsFights.UsedRange.Rows.Count
  if ($fightsRowCount -lt 3 -and -not [string]($wsFights.Cells.Item(2,1).Value2)) {
    $wsFights.Cells.Item(2,$fightsEventIdColIdx).Value2 = 'E001'
    $wsFights.Cells.Item(2,$fightsNameColIdx).Value2 = 'Bout 1'
    $wsFights.Cells.Item(3,$fightsEventIdColIdx).Value2 = 'E002'
    $wsFights.Cells.Item(3,$fightsNameColIdx).Value2 = 'Bout 2'
  }
}

$eventColL = ColLetter $eventIdColIdx
$eventDisplayColL = $null; if ($eventDisplayColIdx) { $eventDisplayColL = (ColLetter $eventDisplayColIdx) }
$fightersEvColL = ColLetter $fightersEventIdColIdx
$fightersNameColL = ColLetter $fightersNameColIdx
if ($fightsEventIdColIdx) { $fightsEvColL = ColLetter $fightsEventIdColIdx }
if ($fightsNameColIdx) { $fightsNameColL = ColLetter $fightsNameColIdx }

# Quoted sheet names for formulas
$eventsSheetName = $wsEvents.Name
$fightersSheetName = $wsFighters.Name
$qEvents = "'" + $eventsSheetName + "'"
$qFighters = "'" + $fightersSheetName + "'"
if ($wsFights) { $fightsSheetName = $wsFights.Name; $qFights = "'" + $fightsSheetName + "'" }

# Ensure Dashboard sheet FIRST (to avoid external link prompts when formulas reference it)
$wsDash = $null
try { $wsDash = $wb.Worksheets.Item('Dashboard') } catch {}
if (-not $wsDash) { $wsDash = $wb.Worksheets.Add(); $wsDash.Name = 'Dashboard' }
$wsDash.Cells.Clear()
$wsDash.Range('B1').Value2 = 'Event'
$wsDash.Range('C1').Value2 = 'Fighter'
$wsDash.Range('D1').Value2 = 'Fight'

# Now ensure Lists sheet and insert formulas referencing Dashboard
$wsLists = $null
try { $wsLists = $wb.Worksheets.Item('Lists') } catch {}
if (-not $wsLists) { $wsLists = $wb.Worksheets.Add(); $wsLists.Name = 'Lists' }
$wsLists.Cells.Clear()
$wsLists.Range('A1').Value2 = 'EventIds'
$wsLists.Range('B1').Value2 = 'FightersByEvent'
$wsLists.Range('C1').Value2 = 'FightsByEvent'
$wsLists.Range('D1').Value2 = 'SelectedEventId'
$wsLists.Range('G1').Value2 = 'EventNames'

$partsEvent = @()
if (-not $safe) {
  $partsEvent = @("=UNIQUE(FILTER(", $qEvents, "!$", $eventColL, ":$", $eventColL, ", ", $qEvents, "!$", $eventColL, ":$", $eventColL, '<>""', "))")
} else {
  $partsEvent = @("=IFERROR(INDEX(", $qEvents, "!$", $eventColL, ":$", $eventColL, ", MATCH(ROW(A1), ", $qEvents, "!$", $eventColL, ":$", $eventColL, ", 0)), \"\")")
}
$formulaEventIds = ($partsEvent -join "")
$wsLists.Range('A2').Formula = $formulaEventIds

$eventNamesAdded = $false
if ($eventDisplayColL) {
  $partsEventNames = @()
  if (-not $safe) {
    $partsEventNames = @("=UNIQUE(FILTER(", $qEvents, "!$", $eventDisplayColL, ":$", $eventDisplayColL, ", ", $qEvents, "!$", $eventDisplayColL, ":$", $eventDisplayColL, '<>""', "))")
  } else {
    $partsEventNames = @("=IFERROR(INDEX(", $qEvents, "!$", $eventDisplayColL, ":$", $eventDisplayColL, ", MATCH(ROW(A1), ", $qEvents, "!$", $eventDisplayColL, ":$", $eventDisplayColL, ", 0)), \"\")")
  }
  $formulaEventNames = ($partsEventNames -join "")
  $wsLists.Range('G2').Formula = $formulaEventNames
  # Selected EventId from chosen Event name
  # More compatible: INDEX/MATCH instead of XLOOKUP (no dynamic arrays)
  $partsSel = @(
    '=IFERROR(INDEX(',
    $qEvents, '!$', $eventColL, ':$', $eventColL,
    ', MATCH(Dashboard!$B$2,',
    $qEvents, '!$', $eventDisplayColL, ':$', $eventDisplayColL,
    ', 0)), "")'
  )
  $formulaSelected = ($partsSel -join "")
  $wsLists.Range('D2').Formula = $formulaSelected
  $eventNamesAdded = $true
} else {
  # Fallback: use direct EventId selection
  $wsLists.Range('D2').Formula = '=Dashboard!$B$2'
}

$partsFighters = @()
if (-not $safe) {
  $partsFighters = @("=UNIQUE(FILTER(", $qFighters, "!$", $fightersNameColL, ":$", $fightersNameColL, ", ", $qFighters, "!$", $fightersEvColL, ":$", $fightersEvColL, '=Lists!$D$2', "))")
} else {
  # Safe: build a helper filtered list using IF / ROW and INDEX without spills
  $partsFighters = @(
    '=IFERROR(INDEX(', $qFighters, '!$', $fightersNameColL, ':$', $fightersNameColL,
    ', SMALL(IF(', $qFighters, '!$', $fightersEvColL, ':$', $fightersEvColL, '=Lists!$D$2, ROW(', $qFighters, '!$', $fightersEvColL, ':$', $fightersEvColL, ')-ROW(', $qFighters, '!$', $fightersEvColL, '$1)+1), ROW(A1))), "")')
}
$formulaFighters = ($partsFighters -join "")
$wsLists.Range('B2').FormulaArray = $formulaFighters

if ($wsFights -and $fightsEventIdColIdx -and $fightsNameColIdx) {
  $partsFights = @()
  if (-not $safe) {
    $partsFights = @("=UNIQUE(FILTER(", $qFights, "!$", $fightsNameColL, ":$", $fightsNameColL, ", ", $qFights, "!$", $fightsEvColL, ":$", $fightsEvColL, '=Lists!$D$2', "))")
    $formulaFights = ($partsFights -join "")
    $wsLists.Range('C2').Formula = $formulaFights
  } else {
    $partsFights = @(
      '=IFERROR(INDEX(', $qFights, '!$', $fightsNameColL, ':$', $fightsNameColL,
      ', SMALL(IF(', $qFights, '!$', $fightsEvColL, ':$', $fightsEvColL, '=Lists!$D$2, ROW(', $qFights, '!$', $fightsEvColL, ':$', $fightsEvColL, ')-ROW(', $qFights, '!$', $fightsEvColL, '$1)+1), ROW(A1))), "")')
    $formulaFights = ($partsFights -join "")
    $wsLists.Range('C2').FormulaArray = $formulaFights
  }
}

# Data validation dropdowns
$xlValidateList = 3
$xlBetween = 1
$dv1 = $wsDash.Range('B2').Validation
# Use Event Names if available; otherwise fallback to EventIds
try { $wsDash.Range('B1').Value2 = if ($eventNamesAdded) { 'Event' } else { 'EventId' } } catch {}
if ($eventNamesAdded) {
  $dv1.Delete(); $dv1.Add($xlValidateList, $xlBetween, 1, '=Lists!$G$2#')
} else {
  $dv1.Delete(); $dv1.Add($xlValidateList, $xlBetween, 1, '=Lists!$A$2#')
}
$dv2 = $wsDash.Range('C2').Validation
$dv2.Delete(); $dv2.Add($xlValidateList, $xlBetween, 1, '=Lists!$B$2#')
try {
  $dv3 = $wsDash.Range('D2').Validation
  $dv3.Delete(); $dv3.Add($xlValidateList, $xlBetween, 1, '=Lists!$C$2#')
} catch {}

# Spill selected fighter headers and values for the chosen fighter (by Name + EventId)
# Determine fighters table width to use entire row references
$fightersLastColIdx = $wsFighters.UsedRange.Columns.Count
if ($fightersLastColIdx -lt $fightersNameColIdx) { $fightersLastColIdx = $fightersNameColIdx }
if ($fightersLastColIdx -lt $fightersEventIdColIdx) { $fightersLastColIdx = $fightersEventIdColIdx }
$fightersLastColL = ColLetter $fightersLastColIdx

$wsLists.Range('H1').Value2 = 'SelectedFighter'
# Headers row (spills across columns to used range)
$partsSelHdr = @('=', $qFighters, '!$A$1:$', $fightersLastColL, '$1')
$formulaSelHdr = ($partsSelHdr -join '')
$wsLists.Range('H2').Formula = $formulaSelHdr

# Selected fighter row (spills across columns to used range)
$partsSelRow = @(
  '=IFERROR(XLOOKUP(Dashboard!$C$2&Lists!$D$2, ',
  $qFighters, '!$', $fightersNameColL, ':$', $fightersNameColL,
  ' & ',
  $qFighters, '!$', $fightersEvColL, ':$', $fightersEvColL,
  ', ',
  $qFighters, '!$A:$', $fightersLastColL, ', ""), "")'
)
$formulaSelRow = ($partsSelRow -join '')
$wsLists.Range('H3').Formula = $formulaSelRow

# Dashboard-friendly label/value view using TRANSPOSE of the spill ranges
$wsDash.Range('A4').Value2 = 'Selected Fighter Stats'
$wsDash.Range('A5').Value2 = 'Field'
$wsDash.Range('B5').Value2 = 'Value'
$wsDash.Range('A6').Formula = '=TRANSPOSE(Lists!$H$2#)'
$wsDash.Range('B6').Formula = '=TRANSPOSE(Lists!$H$3#)'

# Grouped sections for key stats
$wsDash.Range('D5').Value2 = 'STRIKING OFFENSE'
$wsDash.Range('G5').Value2 = 'STRIKING DEFENSE'
$wsDash.Range('J5').Value2 = 'GRAPPLING OFFENSE'
$wsDash.Range('M5').Value2 = 'GRAPPLING DEFENSE'

$strikeOff = @('TSSL','TSSA','STRACC%','SSL/M','SSA/M','DSL/M','DSA/M','KD/15mins')
$strikeDef = @('TSSA','TSSAT','STRDEF%','TSSA/M','TSSAT/M','DSSA/M','DSSAT/M','KDA/15mins')
$grapOff   = @('TTD','TTDA','TD%','TDAVG','SUBAVG','CNTL','CNTL%')
$grapDef   = @('TTD','TTDA','TDDEF%','TDAAVG','SUBAVG','CNTLA','CNTLA%')

function Write-Group([object[]]$labels, [int]$startRow, [int]$labelCol, [int]$valueCol) {
  for ($i = 0; $i -lt $labels.Count; $i++) {
    $r = $startRow + $i
    $lbl = [string]$labels[$i]
    $wsDash.Cells.Item($r, $labelCol).Value2 = $lbl
    $partsVal = @('=IFERROR(XLOOKUP("', $lbl, '", Lists!$H$2#, Lists!$H$3#, ""), "")')
    $wsDash.Cells.Item($r, $valueCol).Formula = ($partsVal -join '')
  }
}

Write-Group $strikeOff 6 4 5   # D/E columns
Write-Group $strikeDef 6 7 8   # G/H columns
Write-Group $grapOff   6 10 11 # J/K columns
Write-Group $grapDef   6 13 14 # M/N columns

# Freeze panes (split screen feel)
try {
  $xl.ActiveWindow.SplitColumn = 1
  $xl.ActiveWindow.SplitRow = 1
  $xl.ActiveWindow.FreezePanes = $true
} catch {}

# Placeholder pie chart using fighter count for selected event
$wsLists.Range('E1').Value2 = 'Label'
$wsLists.Range('F1').Value2 = 'Value'
# Build label and value formulas using the detected fighters name column
$partsLabels = @("=UNIQUE(", $qFighters, "!$", $fightersNameColL, ":$", $fightersNameColL, ")")
$formulaLabels = ($partsLabels -join "")
$wsLists.Range('E2').Formula = $formulaLabels
$wsLists.Range('F2').Formula = '=IF(LEN(E2#)>0,1,"")'

# Create chart on Dashboard linked to E2# / F2#
$charts = $wsDash.ChartObjects()
if ($charts.Count -gt 0) { $charts.Item(1).Delete() }
$co = $wsDash.ChartObjects().Add(300, 20, 360, 220)
$ch = $co.Chart
$ch.ChartType = 5  # xlPie
$ch.SetSourceData($wsLists.Range('F2').Resize(1000,1))
$ch.SeriesCollection(1).XValues = $wsLists.Range('E2').Resize(1000,1)
$ch.HasTitle = $true
$ch.ChartTitle.Text = 'Fighters (demo)'

$wb.Save()
$wb.Close($true)
$xl.Quit()
[System.Runtime.InteropServices.Marshal]::ReleaseComObject($wsDash) | Out-Null
[System.Runtime.InteropServices.Marshal]::ReleaseComObject($wsLists) | Out-Null
[System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) | Out-Null
[System.Runtime.InteropServices.Marshal]::ReleaseComObject($xl) | Out-Null
'''

    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except subprocess.CalledProcessError as e:
        # Surface PowerShell errors to help diagnose sheet/header issues
        raise RuntimeError(
            "PowerShell dashboard build failed.\n" +
            (f"STDOUT:\n{e.stdout}\n" if e.stdout else "") +
            (f"STDERR:\n{e.stderr}\n" if e.stderr else "")
        ) from e


